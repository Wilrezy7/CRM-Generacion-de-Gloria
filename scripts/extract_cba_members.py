import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

SOURCE = Path(r"C:\Users\Usuario\Desktop\DB Ministerio Juvenil CBA.xlsx")
TARGET = Path(r"C:\Users\Usuario\Documents\Codex\2026-04-25\desarrolla-un-sistema-crm-web-profesional\outputs\cba_members.json")

MONTHS = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
    "mayo": 5,
    "junio": 6,
    "julio": 7,
    "agosto": 8,
    "septiembre": 9,
    "setiembre": 9,
    "octubre": 10,
    "noviembre": 11,
    "diciembre": 12,
}


def parse_birth_date(value):
    if value is None:
      return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    if not text:
        return ""
    try:
        return datetime.fromisoformat(text).date().isoformat()
    except ValueError:
        pass
    parts = text.lower().replace(" de ", " ").split()
    if len(parts) >= 3:
        day = int(parts[0])
        month = MONTHS.get(parts[1], 1)
        year = int(parts[2])
        return datetime(year, month, day).date().isoformat()
    return ""


def normalize_role(value):
    raw = (str(value or "").strip().lower())
    if raw in {"líder", "lider", "co-líder", "colider", "co lider"}:
        return "Lider"
    if raw == "mentor":
        return "Mentor"
    if raw == "diacono":
        return "Diacono"
    return "Miembro"


wb = load_workbook(SOURCE, read_only=True, data_only=True)
ws = wb["Base de Datos"]
rows = list(ws.iter_rows(values_only=True))
records = []

for row in rows[2:]:
    if not any(value is not None and str(value).strip() for value in row):
        continue
    full_name = str(row[1] or "").strip()
    document_id = str(row[2] or "").strip()
    phone = str(row[3] or "").strip()
    birth_date = parse_birth_date(row[4])
    email = str(row[5] or "").strip().lower()
    baptized = str(row[6] or "NO").strip().upper()
    member_role = normalize_role(row[7])
    records.append(
        {
            "fullName": full_name,
            "documentId": document_id,
            "phone": phone,
            "birthDate": birth_date,
            "email": email,
            "baptized": baptized if baptized in {"SI", "NO"} else "NO",
            "memberRole": member_role,
            "status": "activo",
            "notes": "",
        }
    )

TARGET.parent.mkdir(parents=True, exist_ok=True)
TARGET.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")
print(TARGET)
